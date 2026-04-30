"""
Import bulk de últimos services + vencimientos de matafuegos a Firestore.

Lee un Excel con columnas:
  DOMINIOS | KM ULTIMO SERVICE | FECHA ULTIMO | MATAFUEGO CHASIS | MATAFUEGO CABINA

y actualiza la colección VEHICULOS de Firestore con los campos:
  ULTIMO_SERVICE_KM             (num)
  ULTIMO_SERVICE_FECHA          (string ISO YYYY-MM-DD)
  VENCIMIENTO_EXTINTOR_EXTERIOR (string ISO YYYY-MM-DD)  ← MATAFUEGO CHASIS
  VENCIMIENTO_EXTINTOR_CABINA   (string ISO YYYY-MM-DD)  ← MATAFUEGO CABINA

Reglas (definidas con Santi 2026-04-30):
  • Patente que NO existe en VEHICULOS → skip, se reporta al final.
  • Celda VACÍA en el Excel → NO se toca el campo existente en Firestore.
  • Si el campo ya tiene valor en Firestore → se SOBREESCRIBE con el del Excel.
  • Sin AUDITORIA_ACCIONES (es import bulk de datos históricos).
  • Idempotente: si valor en Firestore == valor del Excel, no escribe nada.
  • Patentes con espacios al final (ej. 'AH490YJ ') → strip + upper.
  • FECHA ULTIMO en el futuro → warning en consola (no bloquea).

Uso:
    # Primero dry-run, no escribe nada:
    python scripts/importar_servicios_y_matafuegos.py --dry-run

    # Si el resumen pinta bien, correr de verdad:
    python scripts/importar_servicios_y_matafuegos.py

    # Path custom al Excel (opcional, default: scripts/datos_servicios_matafuegos_2026-04-30.xlsx)
    python scripts/importar_servicios_y_matafuegos.py --excel ruta/al/archivo.xlsx --dry-run

Requiere:
  - serviceAccountKey.json en la raíz del proyecto (en .gitignore).
  - openpyxl instalado: pip install openpyxl firebase-admin
"""

import sys
import argparse
import datetime
from pathlib import Path

import firebase_admin
from firebase_admin import credentials, firestore
import openpyxl


DEFAULT_EXCEL = Path(__file__).parent / "datos_servicios_matafuegos_2026-04-30.xlsx"

# Mapeo Excel column → Firestore field
COLUMNAS = {
    "KM ULTIMO SERVICE": "ULTIMO_SERVICE_KM",
    "FECHA ULTIMO": "ULTIMO_SERVICE_FECHA",
    "MATAFUEGO CHASIS": "VENCIMIENTO_EXTINTOR_EXTERIOR",
    "MATAFUEGO CABINA": "VENCIMIENTO_EXTINTOR_CABINA",
}


def conectar():
    try:
        if not firebase_admin._apps:
            cred = credentials.Certificate("serviceAccountKey.json")
            firebase_admin.initialize_app(cred)
        db = firestore.client()
        print("[OK] Conexion a Firebase OK.")
        return db
    except Exception as e:
        print(f"[ERROR] Error de conexion: {e}")
        sys.exit(1)


def fecha_a_iso(value):
    """Convierte un datetime/date de openpyxl a string ISO YYYY-MM-DD.
    Si la celda es None o no es una fecha, devuelve None."""
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")
    # A veces el Excel viene con fechas como string ya formateadas — tolerarlas.
    s = str(value).strip()
    if not s:
        return None
    # Best effort: si ya parece YYYY-MM-DD, devolver tal cual.
    try:
        parsed = datetime.datetime.strptime(s[:10], "%Y-%m-%d")
        return parsed.strftime("%Y-%m-%d")
    except ValueError:
        return None


def km_a_num(value):
    """Convierte el KM del Excel a float. None si vacío."""
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def leer_excel(path: Path):
    """Devuelve lista de dicts con los datos del Excel.
    Cada dict tiene: dominio, km, fecha_service, mat_chasis, mat_cabina, fila."""
    print(f"\nLeyendo Excel: {path}")
    if not path.exists():
        print(f"[ERROR] Archivo no encontrado: {path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Buscar encabezados en la primera fila.
    headers = [cell.value for cell in ws[1]]
    try:
        idx_dominio = headers.index("DOMINIOS")
        idx_km = headers.index("KM ULTIMO SERVICE")
        idx_fecha = headers.index("FECHA ULTIMO")
        idx_chasis = headers.index("MATAFUEGO CHASIS")
        idx_cabina = headers.index("MATAFUEGO CABINA")
    except ValueError as e:
        print(f"[ERROR] Falta columna esperada en el Excel: {e}")
        print(f"   Headers encontrados: {headers}")
        sys.exit(1)

    filas = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        dominio_raw = row[idx_dominio]
        if not dominio_raw:
            continue  # fila vacía
        dominio = str(dominio_raw).strip().upper()
        if not dominio:
            continue
        filas.append({
            "fila": i,
            "dominio": dominio,
            "km": km_a_num(row[idx_km]),
            "fecha_service": fecha_a_iso(row[idx_fecha]),
            "mat_chasis": fecha_a_iso(row[idx_chasis]),
            "mat_cabina": fecha_a_iso(row[idx_cabina]),
        })

    print(f"[OK] {len(filas)} filas leidas del Excel.")
    return filas


def calcular_updates(actual: dict, deseado: dict) -> dict:
    """Devuelve solo los campos que cambian (idempotencia).
    `deseado` es el dict candidato a aplicar. Si un campo del deseado es None,
    significa 'no tocar el existente' (celda vacia en el Excel)."""
    updates = {}
    for campo, val_nuevo in deseado.items():
        if val_nuevo is None:
            continue  # celda vacia → no tocamos
        val_actual = actual.get(campo)

        # Comparacion robusta: KM como float, fechas como string ISO.
        if campo == "ULTIMO_SERVICE_KM":
            actual_num = None
            if val_actual is not None:
                try:
                    actual_num = float(val_actual)
                except (TypeError, ValueError):
                    actual_num = None
            if actual_num != val_nuevo:
                updates[campo] = val_nuevo
        else:
            actual_str = str(val_actual) if val_actual is not None else ""
            if actual_str != val_nuevo:
                updates[campo] = val_nuevo
    return updates


def procesar(db, filas, dry_run: bool):
    hoy = datetime.date.today()
    actualizadas = 0
    sin_cambios = 0
    no_existen = []
    futuras = []
    cambios_log = []

    for f in filas:
        dominio = f["dominio"]
        ref = db.collection("VEHICULOS").document(dominio)
        snap = ref.get()
        if not snap.exists:
            no_existen.append(dominio)
            continue

        # Warning de fecha futura (mas de 30 dias)
        if f["fecha_service"]:
            try:
                fs = datetime.datetime.strptime(f["fecha_service"], "%Y-%m-%d").date()
                if (fs - hoy).days > 30:
                    futuras.append((dominio, f["fecha_service"]))
            except ValueError:
                pass

        actual = snap.to_dict() or {}
        deseado = {
            "ULTIMO_SERVICE_KM": f["km"],
            "ULTIMO_SERVICE_FECHA": f["fecha_service"],
            "VENCIMIENTO_EXTINTOR_EXTERIOR": f["mat_chasis"],
            "VENCIMIENTO_EXTINTOR_CABINA": f["mat_cabina"],
        }
        updates = calcular_updates(actual, deseado)

        if not updates:
            sin_cambios += 1
            continue

        actualizadas += 1
        accion = "DRY-RUN" if dry_run else "OK"
        cambios = []
        for k, v in updates.items():
            previo = actual.get(k, "(vacio)")
            cambios.append(f"{k}: {previo!r} -> {v!r}")
        cambios_log.append(f"  [{accion}] VEHICULOS/{dominio}\n      " + "\n      ".join(cambios))

        if not dry_run:
            ref.update(updates)

    # Reporte
    print("\n" + "=" * 70)
    print("DETALLE DE CAMBIOS")
    print("=" * 70)
    if cambios_log:
        for linea in cambios_log:
            print(linea)
    else:
        print("  (ninguno - todos los datos del Excel ya coinciden con Firestore)")

    if futuras:
        print("\n" + "=" * 70)
        print("WARNINGS: FECHA ULTIMO en el futuro (>30 dias) - revisar Excel")
        print("=" * 70)
        for dom, fecha in futuras:
            print(f"  {dom}  fecha_service={fecha}")

    if no_existen:
        print("\n" + "=" * 70)
        print(f"PATENTES DEL EXCEL QUE NO EXISTEN EN VEHICULOS ({len(no_existen)})")
        print("=" * 70)
        print("Estas se SKIPEARON. Si las queres, cargalas manualmente desde la app:")
        for dom in no_existen:
            print(f"  - {dom}")

    print("\n" + "=" * 70)
    print("RESUMEN")
    print("=" * 70)
    print(f"  Total filas Excel:         {len(filas)}")
    print(f"  Actualizadas:              {actualizadas}")
    print(f"  Sin cambios (idempotente): {sin_cambios}")
    print(f"  Patentes no encontradas:   {len(no_existen)}")
    print(f"  Warnings fecha futura:     {len(futuras)}")
    if dry_run:
        print("\n  >> MODO DRY-RUN - no se escribio nada. Si todo OK, corre sin --dry-run.")
    else:
        print("\n  >> MODO REAL - cambios aplicados a Firestore.")


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--dry-run", action="store_true",
                        help="No escribe en Firestore, solo muestra que haria.")
    parser.add_argument("--excel", default=str(DEFAULT_EXCEL),
                        help=f"Ruta al Excel (default: {DEFAULT_EXCEL.name})")
    args = parser.parse_args()

    if args.dry_run:
        print("[DRY-RUN] Modo dry-run - no se va a escribir nada.\n")
    else:
        print("[REAL] Modo real - se van a modificar documentos en Firestore.\n")
        confirm = input("Confirmas? (escribi SI para continuar): ").strip()
        if confirm != "SI":
            print("Cancelado.")
            sys.exit(0)

    db = conectar()
    filas = leer_excel(Path(args.excel))
    procesar(db, filas, args.dry_run)

    print("\n--- PROCESO FINALIZADO ---")


if __name__ == "__main__":
    main()
